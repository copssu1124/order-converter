# -*- coding: utf-8 -*-
"""아이스앤팩 주문서 변환기 — 엔진.
   원본 주문서 → ①변환(메인 엔진 재사용, 계산) → 형식 맞춤(상품명 *수량·택배사 정렬)
   → ②재변환(매칭표 '수정데이터'로 '수정' 행을 발주용 포장으로 해소, 낱개 분리) → 최종.
   ※ 메인 프로그램(gui.py·step3_convert.py)은 건드리지 않고 그대로 재사용."""
import os
import re
import tempfile
import pandas as pd

import step3_convert as engine   # 1차 변환 계산 재사용

# 택배사 정렬 순서(샘플 기준). 목록에 없는 택배사는 뒤로.
택배사순서 = ['대신', '대신낱개', '대신택배', '로젠', '천일', '수동확인',
              '씨제이', '위플', '원준', '카몬드', '올담', '용차', '수정']

C_상품, C_택배사, C_등급 = '상품명', '택배사', '택배등급'
C_합계, C_배송비, C_정산, C_수량 = '배송비합계', '배송비', '정산예상금액', '수량'


def _star_qty(name):
    """수정품명 끝의 '*N'에서 N(=발주 수량)을 뽑음. 없으면 None."""
    s = '' if name is None else str(name)
    m = re.search(r'\*\s*(\d+)\s*$', s)
    return int(m.group(1)) if m else None


def _fix_zip(v):
    """우편번호 앞자리 0 보존 — 숫자로 읽혀 0이 날아간 경우 5자리로 복원(예: 6748→06748)."""
    s = _norm(v)
    if s.isdigit() and 1 <= len(s) <= 4:
        return s.zfill(5)
    return v


def _norm(v):
    if pd.isna(v):
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _recon_name(상품명, 수량):
    """엔진이 떼어낸 '*수량'을 상품명 끝에 복원(발주 표기·수정데이터 조회 키용)."""
    s = '' if pd.isna(상품명) else str(상품명).strip()
    if s == '' or '수동확인' in s:
        return s
    if re.search(r'\*\s*\d+\s*$', s):        # 이미 있으면 그대로(대신낱개 등)
        return s
    q = _norm(수량)
    return f"{s}*{q}" if q else s


def _load_susjung(susjung_file):
    """수정데이터 파일/시트 → {기존품명: (묶음, 낱개)}.
       '기존품명' 헤더를 찾아 그 오른쪽 8칸을 상대 위치로 읽음(별도 파일/탭 위치 무관).
       배치: 기존품명 | 수정품명-1 | 운임 | 등급 | 택배사 | 수정품명-2 | 운임 | 등급 | 택배사.
       묶음/낱개 = (수정품명, 운임, 등급, 택배사). 낱개(수정품명-2) 없으면 None."""
    xls = pd.ExcelFile(susjung_file)
    sheet = "수정데이터" if "수정데이터" in xls.sheet_names else xls.sheet_names[0]
    d = xls.parse(sheet, header=None, dtype=object)
    xls.close()
    # '기존품명' 헤더 위치 찾기(상단 10행 내)
    hr = hc = None
    for r in range(min(10, len(d))):
        for c in range(d.shape[1]):
            if _norm(d.iloc[r, c]) == '기존품명':
                hr, hc = r, c
                break
        if hr is not None:
            break
    if hr is None:
        raise RuntimeError("수정데이터에서 '기존품명' 열을 못 찾았어요. "
                           "(3번째 칸에 '수정데이터' 표가 든 엑셀을 넣어주세요)")
    lut = {}
    for r in range(hr + 1, len(d)):
        key = _norm(d.iloc[r, hc])
        if not key:
            continue

        def g(off):
            cc = hc + off
            return d.iloc[r, cc] if cc < d.shape[1] else None

        묶음 = (g(1), g(2), g(3), g(4))
        낱개 = (g(5), g(6), g(7), g(8)) if _norm(g(5)) else None
        if key not in lut:
            lut[key] = (묶음, 낱개)
    return lut


def _apply_colors(path, kinds):
    """행 종류별 배경색 규칙 — 수동확인=연빨강, 수정 해소=연보라, 정상=흰색, 헤더=연회색."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    색 = {'수동확인': 'FFFFC7CE', '수정': 'FFE6CCF5'}      # 연빨강 / 연보라
    hdr = PatternFill('solid', fgColor='FFF0F0F0')
    try:
        wb = load_workbook(path)
        ws = wb.active
        ncol = ws.max_column
        for c in range(1, ncol + 1):
            ws.cell(1, c).fill = hdr                       # 헤더행
        for i, k in enumerate(kinds):
            rgb = 색.get(k)
            if not rgb:
                continue
            fill = PatternFill('solid', fgColor=rgb)
            r = i + 2                                       # 1행=헤더 → 데이터는 2행부터
            for c in range(1, ncol + 1):
                ws.cell(r, c).fill = fill
        wb.save(path)
    except Exception:
        pass


def 이차재변환_제자리(path, susjung_file, log=print):
    """이미 색상 입은 1차 결과 파일(convert_v2 산출물)에 2차 재변환을 '제자리' 적용.
       ★ 기존 색상은 그대로 보존. 낱개 분리 행만 원본 수정행의 값·스타일을 복사해 끝에 추가.
       - 수정 행(택배사='수정'): 상품명/합계/등급/택배사/수량 교체, 발송일시 99999 제거. (색=기존 연보라 유지)
       - 낱개(수정품명-2): 원본 수정행 복제 → 값 교체 + 정산·배송비 비움. (색=복사됨)
       - 수정데이터에 없으면: 택배사='수동확인'으로만 표시.
       반환: (해소건수, 미매칭리스트[(상품, 수령자)])."""
    from openpyxl import load_workbook
    from copy import copy
    lut = _load_susjung(susjung_file)
    wb = load_workbook(path)
    ws = wb.active
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            hdr.setdefault(str(v).strip(), c)

    def C(name):
        return hdr.get(name)

    c상품, c택배, c등급, c합계 = C('상품명'), C('택배사'), C('택배등급'), C('배송비합계')
    c수량, c정산, c배송비, c발송, c수령 = C('수량'), C('정산예상금액'), C('배송비'), C('발송일시'), C('수령자')
    if not (c상품 and c택배 and c합계):
        raise RuntimeError("결과 파일에서 상품명/택배사/배송비합계 열을 못 찾았어요.")

    def val(r, c):
        if not c:
            return ''
        x = ws.cell(r, c).value
        return '' if x is None else str(x).strip()

    해소, 미매칭, tails = 0, [], []
    for r in range(2, ws.max_row + 1):
        if val(r, c택배) == '수정':
            hit = lut.get(val(r, c상품))
            if hit:
                묶음, 낱개 = hit
                ws.cell(r, c상품).value = 묶음[0]
                ws.cell(r, c합계).value = 묶음[1]
                if c등급:
                    ws.cell(r, c등급).value = 묶음[2]
                ws.cell(r, c택배).value = 묶음[3]
                q = _star_qty(묶음[0])
                if q is not None and c수량:
                    ws.cell(r, c수량).value = q
                if c발송 and val(r, c발송) == '99999':
                    ws.cell(r, c발송).value = None
                해소 += 1
                if 낱개:
                    tails.append((r, 낱개))
            else:
                ws.cell(r, c택배).value = '수동확인'
                미매칭.append((val(r, c상품), val(r, c수령)))

    nr = ws.max_row
    for srcr, 낱개 in tails:                      # 낱개 행을 맨 끝에 추가(원본 수정행 스타일 복사)
        nr += 1
        for c in range(1, ws.max_column + 1):
            s = ws.cell(srcr, c)
            d = ws.cell(nr, c)
            d.value = s.value
            if s.has_style:
                d._style = copy(s._style)
        ws.cell(nr, c상품).value = 낱개[0]
        ws.cell(nr, c합계).value = 낱개[1]
        if c등급:
            ws.cell(nr, c등급).value = 낱개[2]
        ws.cell(nr, c택배).value = 낱개[3]
        q = _star_qty(낱개[0])
        if q is not None and c수량:
            ws.cell(nr, c수량).value = q
        if c정산:
            ws.cell(nr, c정산).value = None
        if c배송비:
            ws.cell(nr, c배송비).value = None
        if c발송 and val(nr, c발송) == '99999':
            ws.cell(nr, c발송).value = None
    wb.save(path)
    log("② 2차 재변환: 수정 %d건 해소 · 낱개 %d행 추가 · 미매칭 %d건" % (해소, len(tails), len(미매칭)))
    return 해소, 미매칭


def 변환(input_file, mapping_file, susjung_file, output_path, log=print):
    """원본 주문서 → 아이스앤팩 최종 파일 저장.
       input_file  : ① 원래 주문서(.xls)
       mapping_file : ② 1차 변환 엑셀(매칭표)
       susjung_file : ③ 2차 변환 엑셀(수정데이터). 비우면 ②에서 읽음.
       반환 dict: 총, 해소, 수동_1차[(상품,수령자)], 미매칭_2차[(상품,수령자)], 출력."""
    if not input_file or not os.path.exists(input_file):
        raise RuntimeError("① 원래 주문서 파일이 없어요.\n   경로: %s" % input_file)
    if not mapping_file or not os.path.exists(mapping_file):
        raise RuntimeError("② 1차 변환 엑셀(매칭표) 파일이 없어요.\n   경로: %s" % mapping_file)
    if susjung_file and not os.path.exists(susjung_file):
        raise RuntimeError("③ 2차 변환 엑셀(수정데이터) 파일이 없어요.\n   경로: %s" % susjung_file)

    # ── 1차: 메인 엔진으로 계산 ──
    log("① 1차 변환(계산) 중…")
    tmp1 = os.path.join(tempfile.gettempdir(), "_ice_1cha.xlsx")
    try:
        engine.convert(input_file, mapping_file, tmp1, log=lambda *a: None)
    except Exception as e:
        raise RuntimeError("① 1차 변환 실패 — 주문서 형식 또는 1차 변환 엑셀(매칭표)을 확인하세요.\n"
                           "   (시트 누락·파일 열림·형식 오류 등)\n   상세: %s" % e)
    df = pd.read_excel(tmp1, dtype=object)
    try:
        os.remove(tmp1)
    except OSError:
        pass
    for need in (C_상품, C_택배사, '수량'):
        if need not in df.columns:
            raise RuntimeError("1차 변환 결과에 '%s' 열이 없어요. 매칭표/주문서 형식을 확인하세요." % need)

    # 상품명 '*수량' 복원 + 택배사 정렬(그룹 내 순서 유지)
    df[C_상품] = [_recon_name(df.iloc[i][C_상품], df.iloc[i]['수량']) for i in range(len(df))]
    rank = {k: i for i, k in enumerate(택배사순서)}
    df['_r'] = [rank.get(_norm(df.iloc[i][C_택배사]), 999) for i in range(len(df))]
    df = df.sort_values('_r', kind='stable').drop(columns='_r').reset_index(drop=True)

    # ── 2차: '수정' 행을 수정데이터로 해소(낱개는 분리해 끝에 추가) ──
    log("② 2차 변환(수정 해소) 중…")
    src2 = susjung_file or mapping_file
    try:
        lut = _load_susjung(src2)
    except Exception as e:
        if susjung_file:      # 2차 엑셀을 '명시적으로' 넣었는데 형식이 이상 → 오류로 알림
            raise RuntimeError("② 2차 변환 엑셀(수정데이터)을 읽지 못했어요 — 파일/시트 형식을 확인하세요.\n"
                               "   상세: %s" % e)
        lut = {}              # 안 넣었고 매칭표에도 수정데이터 없음 → 2차는 건너뜀(1차만)
        log("ℹ 수정데이터가 없어 2차 재변환은 건너뜁니다(1차 변환 결과).")

    def loc(row):
        r = _norm(row.get('수령자'))
        return "%s (수령자 %s)" % (_norm(row.get(C_상품)) or '상품명 불명', r or '?')

    rows, tails = [], []
    row_kinds, tail_kinds = [], []                 # 색칠용: '정상'/'수정'/'수동확인'
    해소 = 0
    수동_1차, 미매칭_2차 = [], []
    for _, row in df.iterrows():
        tag = _norm(row[C_택배사])
        if tag == '수정':
            hit = lut.get(_norm(row[C_상품]))
            if hit:
                묶음, 낱개 = hit
                a = row.copy()
                a[C_상품], a[C_합계], a[C_등급], a[C_택배사] = 묶음[0], 묶음[1], 묶음[2], 묶음[3]
                qa = _star_qty(묶음[0])                       # 수량 = 수정품명 끝의 *N
                if qa is not None:
                    a[C_수량] = qa
                rows.append(a); row_kinds.append('수정')
                해소 += 1
                if 낱개:
                    b = row.copy()
                    b[C_상품], b[C_합계], b[C_등급], b[C_택배사] = 낱개[0], 낱개[1], 낱개[2], 낱개[3]
                    qb = _star_qty(낱개[0])
                    if qb is not None:
                        b[C_수량] = qb
                    b[C_정산], b[C_배송비] = None, None      # 낱개 줄: 정산·배송비 비움
                    tails.append(b); tail_kinds.append('수정')
            else:
                x = row.copy()
                x[C_택배사] = '수동확인'                     # 수정데이터에 없으면 수동확인
                rows.append(x); row_kinds.append('수동확인')
                미매칭_2차.append(loc(row))
        else:
            rows.append(row)
            row_kinds.append('수동확인' if tag == '수동확인' else '정상')
            if tag == '수동확인':
                수동_1차.append(loc(row))

    out = pd.DataFrame(rows + tails).reset_index(drop=True)
    if '우편번호' in out.columns:                        # 앞자리 0 복원(배송 필수)
        out['우편번호'] = out['우편번호'].map(_fix_zip)
    if '발송일시' in out.columns:                        # 수정 sentinel 99999 제거
        out['발송일시'] = out['발송일시'].map(lambda v: None if _norm(v) == '99999' else v)
    try:
        out.to_excel(output_path, index=False)
    except PermissionError:
        raise RuntimeError("결과 파일을 저장하지 못했어요 — 같은 이름 파일이 열려 있으면 닫고 다시 해주세요.\n"
                           "   경로: %s" % output_path)
    _apply_colors(output_path, row_kinds + tail_kinds)   # 색상 규칙 적용

    # 상세 로그(누락 제품 안내)
    log("─" * 30)
    log("✅ 완료: 총 %d행 · 수정 해소 %d건" % (len(out), 해소))
    if 미매칭_2차:
        log("⚠ [2차 미해결] %d건 — ③ 수정데이터에 없는 상품(추가하면 자동 해소):" % len(미매칭_2차))
        for s in 미매칭_2차:
            log("     · " + s)
    if 수동_1차:
        log("⚠ [수동확인] %d건 — ② 매칭표에 없는 상품(직접 확인/등록 필요):" % len(수동_1차))
        for s in 수동_1차:
            log("     · " + s)
    if not 미매칭_2차 and not 수동_1차:
        log("빠짐없이 전부 변환됐어요. 👍")
    return {'총': len(out), '해소': 해소, '수동_1차': 수동_1차,
            '미매칭_2차': 미매칭_2차, '출력': output_path}
