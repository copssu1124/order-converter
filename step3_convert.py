import sys
import os
import re
import glob
import datetime
import pandas as pd
import unicodedata
import openpyxl
import xlrd
from collections import Counter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# ============================================================
# 주문서 변환 엔진 (GUI/CLI 공용)
#   - 변환 로직(상품명·택배비·천일박스·색·그룹·택배등급)은 그대로 유지
#   - convert()로 호출. print 대신 log 콜백으로 메시지 전달.
#   - 개인/고정 경로 하드코딩 없음. 출력은 입력 파일 폴더에 생성.
# ============================================================

# 매핑 사전 — convert()에서 매핑표 파일을 읽어 채움 (헬퍼 함수가 전역으로 참조)
매핑표 = {}
판매비표 = {}
천일택배비표 = {}
씨제이택배비표 = {}
로젠택배비표 = {}
대신택배표 = {}

# 매핑표 파일 자동 탐색용 패턴 (이름에 '상품명'과 '변경' 포함된 .xlsx)
MAPPING_GLOBS = ["*상품명*택배사*변경*.xlsx", "*상품명*변경*.xlsx", "★*.xlsx"]
INPUT_GLOB = "제이제이 *.xls"


# ─────────────────────────────────────────────────────────
# 묶음수량 계산 — 배송비합계 = 배송비단가 × 묶음수량
#   ★ 반제품 키워드 있으면 → 무조건 1 (단위표시가 있어도 합포장, 최우선)
#   (숫자ea) 또는 (숫자매) 패턴 있으면 개별 배송 → 수량 그대로
#   그 외(낱개 등) 합포장 → 1
# ─────────────────────────────────────────────────────────
def 묶음수량계산(회사상품명, 수량):
    if '반제품' in 회사상품명:        # 최우선 — ea/매 표시가 있어도 ×1
        return 1
    if re.search(r'\(\d+ea\)', 회사상품명, re.IGNORECASE):
        return 수량
    if re.search(r'\(\d+매\)', 회사상품명):
        return 수량
    return 1


# ─────────────────────────────────────────────────────────
# 정규화
# ─────────────────────────────────────────────────────────
def 정규화(s):
    s = ''.join(c for c in s if unicodedata.category(c)[0] != 'C')
    s = s.replace('\xa0', ' ').replace('　', ' ')
    s = re.sub(r' +', ' ', s)
    s = s.strip()
    s = s.lower()
    return s


# ─────────────────────────────────────────────────────────
# 상품명 변환 (기존 로직 그대로)
# ─────────────────────────────────────────────────────────
def 변환시도(키, 수량):
    키_norm = 정규화(키)
    if 키_norm in 매핑표:
        return 매핑표[키_norm] + str(수량), "직접매핑"
    키_esm = re.sub(r'/\s*\d[\d,]*\s*원\s*/\s*\d+\s*개', '', 키_norm).strip()
    if 키_esm != 키_norm and 키_esm in 매핑표:
        return 매핑표[키_esm] + str(수량), "ESM접미사제거"
    키_11 = re.sub(r'-\d+개$', '', 키_norm).strip()
    if 키_11 != 키_norm and 키_11 in 매핑표:
        return 매핑표[키_11] + str(수량), "11번가접미사제거"
    return "수동확인 필요", "매핑실패"


# ─────────────────────────────────────────────────────────
# 판매비 파싱 및 조회
# ─────────────────────────────────────────────────────────
def parse_result(v):
    """'가격*택배사' 형식 -> (택배비, 택배사)"""
    if '*' not in v:
        return v, '?'
    fee_part, carrier_raw = v.split('*', 1)
    fee_part = fee_part.strip()
    carrier_raw = carrier_raw.strip()
    if not carrier_raw:
        carrier = '대신'      # 별표 뒤 빈 순수금액값 → 대신 (이전: 천일)
    elif '대신택배' in carrier_raw:   # '대신' 포함이라 대신낱개·대신보다 먼저 판정
        carrier = '대신택배'
    elif '대신낱개' in carrier_raw:
        carrier = '대신낱개'
    elif '다모아' in carrier_raw:
        carrier = '다모아'
    elif '씨제이' in carrier_raw:
        carrier = '씨제이'
    elif '로젠' in carrier_raw:
        carrier = '로젠'
    elif '원준' in carrier_raw:
        carrier = '원준'
    elif '위플' in carrier_raw:
        carrier = '위플'
    else:
        carrier = carrier_raw
    m = re.match(r'^(\d+)', fee_part)
    fee = m.group(1) if m else (fee_part if fee_part else '-')
    return fee, carrier


def 천일박스파싱(박스값):
    """천일 매핑값 -> (박스리스트, is_total)
       박스리스트: [(박스명, 박스가격int, 박스개수), ...] / None(가격조회실패) / [](박스없음)
       is_total: True=이미 총량(수량-M·이중박스) / False=단일박스(1개당1박스, 호출부에서 ×묶음수량)

    형식:
      '스티로폴박스N'              -> 박스N ×1   (단일박스, is_total=False)
      '스티로폴박스N/수량-M'       -> 박스N ×M   (총량, is_total=True)
      '스티로폴박스N(YEA)/수량-M'  -> 박스N ×M   (총량, is_total=True)
      '스티로폴박스A(YEA)/스티로폴박스B(WEA)' -> 박스A ×1 + 박스B ×1 (이중박스, is_total=True)
      변종: '수량1'(하이픈 없음), '수량*M', 끝의 '*' 등도 흡수
    """
    박스번호들 = re.findall(r'스티로폴박스(\d)', 박스값)
    if not 박스번호들:
        return [], False  # 박스 정보 없음 → 수동확인 대상

    def 가격(N):
        v = 천일택배비표.get(정규화(f'스티로폴박스{N}'))
        return int(v) if v is not None else None

    if len(박스번호들) == 1:
        N = 박스번호들[0]
        m = re.search(r'수량\s*[-*]?\s*(\d+)', 박스값)  # 수량-M / 수량M / 수량*M
        M = int(m.group(1)) if m else 1
        p = 가격(N)
        if p is None:
            return None, False
        # is_total: '수량-M'이 명시되면 이미 주문수량 반영된 총량형.
        #           없으면(단일박스) 1개당 1박스 → 호출부에서 ×묶음수량.
        is_total = m is not None
        return [(f'스티로폴박스{N}', p, M)], is_total

    # 이중(다중)박스 — 각 박스 등장 횟수만큼 개수 (총량형)
    박스리스트 = []
    for N, c in Counter(박스번호들).items():
        p = 가격(N)
        if p is None:
            return None, False
        박스리스트.append((f'스티로폴박스{N}', p, c))
    return 박스리스트, True


def 박스번호목록(박스값):
    """천일 값에서 '스티로폴박스N' 번호만 추출 (개수 제거, 중복제거, 순서유지)."""
    nums = []
    for d in re.findall(r'스티로폴박스(\d)', str(박스값)):
        name = '스티로폴박스' + d
        if name not in nums:
            nums.append(name)
    return ' '.join(nums)


def 택배비조회(상품키):
    """일반 키 우선 조회. 천일 키는 택배등급(박스번호)용으로만 참조.
       반환 dict: 경로('일반'/'천일'/'없음'), carrier, 값, 씨제이등급, 박스등급, 천일박스, 비고
       (배송비합계·단가 계산은 호출부에서 수량/묶음수량으로 수행)"""
    일반키 = 정규화(상품키)
    천일키 = 정규화(상품키 + '천일')
    박스등급 = 박스번호목록(판매비표[천일키]) if 천일키 in 판매비표 else ''

    일반_순수금액 = False
    if 일반키 in 판매비표:                       # 1순위: 일반 키 (형식 정상일 때만)
        값 = 판매비표[일반키]
        fee, carrier = parse_result(값)
        if carrier != '?':                       # '*' 없는 깨진 값은 일반으로 안 씀 → 천일 폴백
            씨제이등급 = fee.upper() if carrier == '씨제이' else None
            로젠등급 = fee.strip().upper() if carrier == '로젠' else None
            대신택배등급 = fee.strip().upper() if carrier == '대신택배' else None
            return {'경로': '일반', 'carrier': carrier, '값': 값,
                    '씨제이등급': 씨제이등급, '로젠등급': 로젠등급, '대신택배등급': 대신택배등급,
                    '박스등급': 박스등급, '천일박스': None, '비고': f'[일반] {값}'}
        # 깨진 일반값이 숫자로 시작(순수금액)이면 합계는 천일로 내되 택배사는 '대신'
        일반_순수금액 = bool(re.match(r'^\s*\d', 값))
    if 천일키 in 판매비표:                       # 2순위: 천일 (일반 없거나 깨짐)
        박스리스트, is_total = 천일박스파싱(판매비표[천일키])
        천일carrier = '대신' if 일반_순수금액 else '천일'
        return {'경로': '천일', 'carrier': 천일carrier, '값': 판매비표[천일키],
                '씨제이등급': None, '로젠등급': None, '대신택배등급': None,
                '박스등급': 박스등급, '천일박스': (박스리스트, is_total),
                '비고': f'[천일] {판매비표[천일키]}'}
    return {'경로': '없음', 'carrier': '수동확인', '값': '', '씨제이등급': None,
            '로젠등급': None, '대신택배등급': None, '박스등급': '',
            '천일박스': None, '비고': '[없음]'}


# ─────────────────────────────────────────────────────────
# 매핑표 / 입력파일 자동 탐색 + 출력경로 생성
# ─────────────────────────────────────────────────────────
def find_mapping_file(folder):
    """폴더에서 매핑표(.xlsx) 자동 탐색. 없으면 None."""
    for pat in MAPPING_GLOBS:
        hits = [f for f in glob.glob(os.path.join(folder, pat))
                if not os.path.basename(f).startswith("변환완료")
                and not os.path.basename(f).startswith("~$")]
        if hits:
            return max(hits, key=os.path.getmtime)
    return None


def auto_detect_input(folder):
    """폴더에서 최신 '제이제이 *.xls' 주문서 자동 선택. 없으면 None."""
    cands = sorted(glob.glob(os.path.join(folder, INPUT_GLOB)),
                   key=os.path.getmtime, reverse=True)
    return cands[0] if cands else None


def make_output_path(input_file):
    """입력 파일 폴더에 '변환완료_주문서_<날짜>.xlsx' 경로 생성."""
    folder = os.path.dirname(os.path.abspath(input_file))
    m = re.search(r'(\d{2})\.(\d{2})', os.path.basename(input_file))
    date = (m.group(1) + m.group(2)) if m else datetime.datetime.now().strftime("%m%d")
    return os.path.join(folder, f"변환완료_주문서_{date}.xlsx")


def _load_mapping(mapping_file, log):
    """매핑표 4개 시트를 읽어 전역 사전 채움 (읽기 전용)."""
    global 매핑표, 판매비표, 천일택배비표, 씨제이택배비표, 로젠택배비표, 대신택배표
    매핑원본 = pd.read_excel(mapping_file, sheet_name="상품명변경", header=None)
    판매비원본 = pd.read_excel(mapping_file, sheet_name="판매비변경", header=None)
    천일택배비원본 = pd.read_excel(mapping_file, sheet_name="천일택배비", header=None)
    씨제이택배비원본 = pd.read_excel(mapping_file, sheet_name="씨제이택배비", header=None)
    try:
        로젠택배비원본 = pd.read_excel(mapping_file, sheet_name="로젠택배비", header=None)
    except Exception:
        로젠택배비원본 = None
    try:
        대신택배원본 = pd.read_excel(mapping_file, sheet_name="대신택배", header=None)
    except Exception:
        대신택배원본 = None

    매핑표 = {}
    for i in range(3, len(매핑원본)):
        원본 = 매핑원본.iloc[i, 7]
        변환 = 매핑원본.iloc[i, 8]
        if pd.notna(원본) and pd.notna(변환):
            매핑표[정규화(str(원본))] = str(변환).strip()

    판매비표 = {}
    for i in range(99, len(판매비원본)):
        키 = 판매비원본.iloc[i, 5]
        값 = 판매비원본.iloc[i, 6]
        if pd.notna(키) and pd.notna(값):
            판매비표[정규화(str(키))] = str(값).strip()

    천일택배비표 = {}
    for i in range(len(천일택배비원본)):
        박스 = 천일택배비원본.iloc[i, 0]
        금액 = 천일택배비원본.iloc[i, 1]
        if pd.notna(박스) and pd.notna(금액):
            천일택배비표[정규화(str(박스))] = str(int(금액))

    씨제이택배비표 = {}
    for i in range(len(씨제이택배비원본)):
        등급 = 씨제이택배비원본.iloc[i, 0]
        금액 = 씨제이택배비원본.iloc[i, 1]
        if pd.notna(등급) and pd.notna(금액):
            씨제이택배비표[str(등급).strip().upper()] = str(int(금액))

    로젠택배비표 = {}                       # 로젠: 씨제이와 동일한 등급 가격표 방식
    if 로젠택배비원본 is not None:
        for i in range(len(로젠택배비원본)):
            등급 = 로젠택배비원본.iloc[i, 0]
            금액 = 로젠택배비원본.iloc[i, 1]
            if pd.notna(등급) and pd.notna(금액):
                로젠택배비표[str(등급).strip().upper()] = str(int(금액))

    대신택배표 = {}                          # 대신택배: 등급 가격표(등급이 숫자 5/6 등)
    if 대신택배원본 is not None:
        for i in range(len(대신택배원본)):
            등급 = 대신택배원본.iloc[i, 0]
            금액 = 대신택배원본.iloc[i, 1]
            if pd.notna(등급) and pd.notna(금액):
                try:
                    키 = str(int(등급))      # 숫자 등급 5,6 → '5','6'
                except (ValueError, TypeError):
                    키 = str(등급).strip().upper()
                대신택배표[키] = str(int(금액))

    log(f"매핑표 로드: 상품명 {len(매핑표)}건 / 판매비 {len(판매비표)}건")


def _read_original_colors(input_file, log):
    """원본 .xls 배경색 읽기 → 출력 Excel 셀 위치로 변환.
       삽입 2개(배송비합계=pos17, 택배등급=pos18) 기준:
         src_col < 17 → excel_col = src_col+1
         src_col >= 17 → excel_col = src_col+3"""
    colors = {}
    try:
        src_wb = xlrd.open_workbook(input_file, formatting_info=True)
        src_ws = src_wb.sheet_by_index(0)
        for xlrd_row in range(src_ws.nrows):
            excel_row_out = xlrd_row + 1
            for src_col in range(src_ws.ncols):
                xf = src_wb.xf_list[src_ws.cell_xf_index(xlrd_row, src_col)]
                bg_idx = xf.background.pattern_colour_index
                if bg_idx not in (0, 64, 65):
                    color = src_wb.colour_map.get(bg_idx)
                    if color and len(color) == 3:
                        r, g, b = color
                        hex_color = f'{r:02X}{g:02X}{b:02X}'
                        excel_col_out = src_col + 1 if src_col < 17 else src_col + 3
                        colors[(excel_row_out, excel_col_out)] = hex_color
        log(f"원본 배경색 셀: {len(colors)}개 보존")
    except Exception as e:
        log(f"[경고] 원본 배경색 읽기 실패 (색상 복원 생략): {e}")
    return colors


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


# ─────────────────────────────────────────────────────────
# 메인 변환 함수
# ─────────────────────────────────────────────────────────
def convert(input_file, mapping_file, output_path=None, log=print, verbose=False):
    """주문서(.xls)를 변환해 새 .xlsx로 저장. 저장된 경로 문자열 반환.
       log: 메시지 콜백 (GUI/콘솔). 실패 시 RuntimeError 발생."""
    if not input_file or not os.path.exists(input_file):
        raise RuntimeError(f"주문서 파일을 찾을 수 없습니다: {input_file}")
    if not mapping_file or not os.path.exists(mapping_file):
        raise RuntimeError(f"매핑표 파일을 찾을 수 없습니다: {mapping_file}")
    if output_path is None:
        output_path = make_output_path(input_file)

    log(f"주문서: {os.path.basename(input_file)}")
    log(f"매핑표: {os.path.basename(mapping_file)}")

    # 1) 파일 로드
    주문서 = pd.read_excel(input_file)
    try:
        _load_mapping(mapping_file, log)
    except Exception as e:
        raise RuntimeError(f"매핑표를 읽을 수 없습니다 (파일이 열려 있거나 시트 누락?): {e}")

    # 2) 원본 색상
    original_excel_colors = _read_original_colors(input_file, log)

    # 3) 메인 루프 — 빈 행에서 중단, 실제 주문건만 처리
    결과목록 = []
    for idx in range(len(주문서)):
        수령자 = 주문서.iloc[idx, 4]
        상품명 = 주문서.iloc[idx, 8]
        if (pd.isna(수령자) or not str(수령자).strip()) or \
           (pd.isna(상품명) or not str(상품명).strip()):
            break

        선택사항 = 주문서.iloc[idx, 7]
        수량_raw = 주문서.iloc[idx, 10]
        수량 = int(수량_raw) if pd.notna(수량_raw) else 1
        키 = str(선택사항).strip() if pd.notna(선택사항) and str(선택사항).strip() \
             else str(상품명).strip()

        상품결과, 방식 = 변환시도(키, 수량)
        회사상품명 = 상품결과.rsplit('*', 1)[0] if '*' in 상품결과 else 상품결과
        묶음수량 = 묶음수량계산(회사상품명, 수량)

        fee = None
        배송비합계 = None
        택배등급 = ''
        if 방식 == "매핑실패":
            택배사_disp = '수동확인'
            is_수동확인 = True
            택배비표시 = '?'
            비고 = '상품명변환실패'
        else:
            조회 = 택배비조회(상품결과)
            경로 = 조회['경로']
            carrier = 조회['carrier']
            비고 = 조회['비고']
            박스등급 = 조회['박스등급']
            씨제이등급 = 조회['씨제이등급']
            로젠등급 = 조회.get('로젠등급')
            대신택배등급 = 조회.get('대신택배등급')

            if 경로 == '일반':
                값 = 조회['값']
                is_수동확인 = (carrier == '?')
                # ── 배송비합계/단가 ──
                #   다모아·원준·위플 = 빈칸
                #   씨제이·로젠·대신택배(등급) = 등급가×수량, 단가 = 1개 등급가격
                #   순수 금액 = 그대로(총액), 단가 = 합계
                if carrier in ('다모아', '원준', '위플'):
                    배송비합계 = None
                    fee = None
                elif carrier == '대신택배' and 대신택배등급 and 대신택배등급 in 대신택배표:
                    단가 = int(대신택배표[대신택배등급])
                    배송비합계 = 단가 * 묶음수량
                    fee = 단가
                elif carrier == '로젠' and 로젠등급 and 로젠등급 in 로젠택배비표:
                    단가 = int(로젠택배비표[로젠등급])
                    배송비합계 = 단가 * 묶음수량
                    fee = 단가
                elif 씨제이등급 and 씨제이등급 in 씨제이택배비표:
                    단가 = int(씨제이택배비표[씨제이등급])
                    배송비합계 = 단가 * 묶음수량
                    fee = 단가
                else:
                    m = re.match(r'^(\d+)', 값)
                    if m:                              # 숫자로 시작 → 이미 총액
                        배송비합계 = int(m.group(1))
                        fee = 배송비합계               # 단가 = 합계 (나누지 않음)
                    else:
                        배송비합계 = None
                        fee = None
                # 택배등급: 천일박스 > 씨제이 > 로젠 > 대신택배
                택배등급 = 박스등급 if 박스등급 else (씨제이등급 or 로젠등급 or 대신택배등급 or '')
                택배사_disp = '수동확인' if is_수동확인 else carrier
                택배비표시 = (str(fee) if fee is not None
                              else (carrier if carrier and carrier != '?' else ''))
            elif 경로 == '천일':                       # 일반 없거나 깨짐 → 천일 박스로 합계
                is_수동확인 = False
                택배사_disp = carrier                   # '천일' 또는 '대신'(일반 순수금액 폴백)
                박스리스트, is_total = 조회['천일박스']
                if 박스리스트:
                    총액 = sum(p * c for _, p, c in 박스리스트)
                    배송비합계 = 총액 if is_total else 총액 * 묶음수량
                    fee = 배송비합계                    # 단가 = 합계 (나누지 않음)
                    택배비표시 = str(fee)
                else:
                    배송비합계 = None
                    택배비표시 = '수동확인'
                택배등급 = 박스등급 or ('조회실패' if 박스리스트 is None else '')
            else:                                       # 없음 → 수동확인
                is_수동확인 = True
                택배사_disp = '수동확인'
                택배비표시 = '수동확인'

        결과목록.append({
            '행': idx, '상품결과': 상품결과, '회사상품명': 회사상품명,
            '택배사': 택배사_disp, '택배비': 택배비표시, 'fee': fee,
            '수량': 수량, '묶음수량': 묶음수량, '배송비합계': 배송비합계,
            '택배등급': 택배등급, '비고': 비고, 'is_수동확인': is_수동확인,
        })

    if not 결과목록:
        raise RuntimeError("처리할 주문건이 없습니다 (빈 주문서이거나 형식이 다릅니다).")

    수동확인수 = sum(1 for r in 결과목록 if r['is_수동확인'])
    log(f"변환 완료: 총 {len(결과목록)}건 (수동확인 {수동확인수}건)")

    if verbose:
        W = 95
        log("=" * W)
        log(f"{'행':>3} | {'회사상품명*수량':<36} | {'택배사':<8} | {'택배비':>7} | 비고")
        log("-" * W)
        for r in 결과목록:
            상품 = r['상품결과']
            상품표시 = (상품[:34] + '..') if len(상품) > 36 else 상품
            log(f"{r['행']:>3} | {상품표시:<36} | {r['택배사']:<8} | {r['택배비']:>7} | {r['비고']}")
        log("=" * W)

    # 4) Excel 데이터 작성 — 원본 NaN 보존, 가짜 값 금지 (저장값은 결과목록에서만)
    출력df = 주문서.iloc[:len(결과목록)].copy()
    출력df[출력df.columns[17]] = None  # 배송번호 비우기 (거짓 데이터 제거)
    출력df[출력df.columns[8]]  = [r['회사상품명'] for r in 결과목록]
    출력df[출력df.columns[11]] = [r['택배사'] for r in 결과목록]
    출력df[출력df.columns[16]] = [r['fee'] for r in 결과목록]
    출력df.insert(17, '배송비합계', [r['배송비합계'] for r in 결과목록])
    출력df.insert(18, '택배등급', [r['택배등급'] for r in 결과목록])

    # 저장 (열려 있으면 시각 붙인 새 파일로 폴백)
    try:
        출력df.to_excel(output_path, index=False)
    except PermissionError:
        base, ext = os.path.splitext(output_path)
        output_path = f"{base}_{datetime.datetime.now().strftime('%H%M%S')}{ext}"
        log(f"[알림] 기존 출력 파일이 열려 있어 새 이름으로 저장합니다.")
        출력df.to_excel(output_path, index=False)

    # 5) 색상 + 테두리 + 병합
    COL_택배사_FILL = _fill('DDEBF7')
    DAMOA_FILL     = _fill('FFC000')
    MANUAL_FILL    = _fill('FF0000')
    COL_배송비_FILL = _fill('E2EFDA')
    SUBTOTAL_BASE  = _fill('FFCCCC')
    SUBTOTAL_MULTI = _fill('FF9999')
    YELLOW         = _fill('FFFF00')

    EXCEL_COL_택배사 = 12
    EXCEL_COL_배송비 = 17
    EXCEL_COL_합계  = 18
    LAST_COL = len(출력df.columns)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    n_rows = ws.max_row

    for (excel_row, excel_col), hex_color in original_excel_colors.items():
        if excel_row <= n_rows:
            ws.cell(row=excel_row, column=excel_col).fill = _fill(hex_color)

    ws.cell(row=1, column=EXCEL_COL_합계).fill = SUBTOTAL_BASE
    ws.cell(row=1, column=EXCEL_COL_합계).font = Font(bold=True)

    for i, r in enumerate(결과목록):
        excel_row = i + 2
        빔 = r['fee'] is None
        if r['택배사'] == '다모아':
            L_fill = DAMOA_FILL
        elif r['택배사'] == '수동확인':
            L_fill = MANUAL_FILL
        else:
            L_fill = COL_택배사_FILL
        ws.cell(row=excel_row, column=EXCEL_COL_택배사).fill = L_fill
        ws.cell(row=excel_row, column=EXCEL_COL_배송비).fill = YELLOW if 빔 else COL_배송비_FILL
        합계_cell = ws.cell(row=excel_row, column=EXCEL_COL_합계)
        합계_cell.font = Font(bold=True)
        if 빔:
            합계_cell.fill = YELLOW
        elif r['묶음수량'] >= 2:
            합계_cell.fill = SUBTOTAL_MULTI
        else:
            합계_cell.fill = SUBTOTAL_BASE

    # 주소 중복 표시 — 주소가 완전히 동일한 줄이 2개 이상이면 그 주소 셀을 빨강
    ADDR_FILL = _fill('FFC7CE')      # 연한 빨강(중복 강조, 글자 보임)
    EXCEL_COL_주소 = 7               # 주문서 주소(col6) → 출력 7열
    addr_list = [str(주문서.iloc[r['행'], 6]).strip() for r in 결과목록]
    addr_cnt = Counter(a for a in addr_list if a and a.lower() != 'nan')
    주소중복행 = 0
    for i, a in enumerate(addr_list):
        if a and addr_cnt.get(a, 0) >= 2:
            ws.cell(row=i + 2, column=EXCEL_COL_주소).fill = ADDR_FILL
            주소중복행 += 1

    wb.save(output_path)

    # 6) 요약
    빈칸수 = sum(1 for r in 결과목록 if r['fee'] is None)
    다모아수 = sum(1 for r in 결과목록 if r['택배사'] == '다모아')
    등급채움 = sum(1 for r in 결과목록 if r['택배등급'] not in ('', '조회실패'))
    log(f"확인필요(노랑 빈칸) {빈칸수}건 / 다모아(주황) {다모아수}건 / 택배등급 {등급채움}건")
    log(f"주소 중복(빨강) {주소중복행}행")
    return output_path


# ─────────────────────────────────────────────────────────
# 오류 사후 분석 — 변환 로직 미수정. 끝난 결과에서 '수동확인(빨강)' 행만
# 골라 사유를 판별하고, 결과 엑셀 맨 끝에 '비고' 열을 추가(기존 열 미변경).
#   상품명변경에 옵션명 없음        → "상품명 미등록"
#   상품명은 있으나 판매비 키 없음   → "택배비 미등록: <회사상품명*수량>"
#   그 외                          → "확인 필요: ..."
# 반환: [(엑셀행, 표시상품명, 사유), ...]
# ─────────────────────────────────────────────────────────
def 분석_수동확인사유(input_file, mapping_file, output_file, log=print):
    _load_mapping(mapping_file, lambda *a: None)   # 변환 때와 동일 매핑 로드(읽기)
    주문서 = pd.read_excel(input_file)

    사유목록 = []
    for idx in range(len(주문서)):
        수령자 = 주문서.iloc[idx, 4]
        상품명 = 주문서.iloc[idx, 8]
        if (pd.isna(수령자) or not str(수령자).strip()) or \
           (pd.isna(상품명) or not str(상품명).strip()):
            break
        선택사항 = 주문서.iloc[idx, 7]
        수량_raw = 주문서.iloc[idx, 10]
        수량 = int(수량_raw) if pd.notna(수량_raw) else 1
        키 = str(선택사항).strip() if pd.notna(선택사항) and str(선택사항).strip() \
             else str(상품명).strip()

        상품결과, 방식 = 변환시도(키, 수량)
        excel_row = idx + 2
        if 방식 == "매핑실패":
            사유목록.append((excel_row, str(상품명),
                           "상품명 미등록 (상품명변경 탭에 옵션명 없음)"))
            continue
        조회 = 택배비조회(상품결과)
        if 조회['경로'] == '없음':       # 천일/일반 어디에도 키 없음
            사유목록.append((excel_row, 상품결과,
                           "택배비 미등록: %s (판매비변경 탭에 키 없음)" % 상품결과))
        elif 조회['carrier'] == '?':     # 값 형식 이상 등
            사유목록.append((excel_row, 상품결과,
                           "확인 필요: %s (값 형식 확인)" % 상품결과))

    # 결과 엑셀 맨 끝에 '비고' 열 추가 (기존 열·서식 미변경)
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    비고열 = ws.max_column + 1
    ws.cell(row=1, column=비고열, value='비고')
    ws.cell(row=1, column=비고열).font = Font(bold=True)
    for excel_row, _disp, 사유 in 사유목록:
        ws.cell(row=excel_row, column=비고열, value=사유)
    try:
        wb.save(output_file)
    except PermissionError:
        log("[알림] 결과 파일이 열려 있어 비고열을 추가하지 못했습니다(목록은 아래 표시).")
    return 사유목록


# ═════════════════════════════════════════════════════════
# 신규: 헤더 인식 + 양식 감지 + 어댑터 + 발화주명/파일명/상품명*수량
#   ※ 기존 convert()·매칭·배송비·색상·그룹 함수는 일절 변경하지 않음.
#      스마트스토어는 제이제이 컬럼배치로 '어댑터' 후 convert() 그대로 호출.
# ═════════════════════════════════════════════════════════

# 제이제이 원본 주문서의 표준 컬럼 순서 (어댑터 출력 틀)
JEJE_HEADERS = [
    'Unnamed: 0', '별칭(쇼핑몰계정)', '수령자휴대전화', '수령자전화번호', '수령자',
    '우편번호', '주소', '선택사항', '상품명', '배송구분', '수량', '택배사', '송장번호',
    '배송메모', '상품주문번호(스마트스토어)', '정산예상금액', '배송비', '배송번호', '배송방법',
    '단가', '옵션추가금액', '총주문금액', '마켓수수료금액', '실수수료율', '결제일시', '주문자ID',
    '구매자', '구매자휴대전화', '주문번호', '도서산간', '발송예정일', '할인금액', '주문고유코드',
    '공급금액', '주문일시', '상품번호', '수수료율', '발송일시', '상품URL', '정산추가정보',
    '실결제금액', '송장번호입력일시(로컬)', '톡톡하기', 'CS메모',
]


def _hdr(s):
    return '' if s is None else str(s).replace(' ', '').lower()


def detect_fields(df):
    """헤더 이름(부분일치)으로 필수 필드의 컬럼 인덱스를 찾는다."""
    H = [_hdr(c) for c in df.columns]
    found = {}

    def pick(field, test):
        for i, h in enumerate(H):
            if test(h):
                found[field] = i
                return

    pick('상품명', lambda h: ('상품명' in h or '제품명' in h) and '주문번호' not in h)
    pick('옵션', lambda h: '선택사항' in h or '옵션정보' in h or '옵션명' in h or '옵션' in h)
    pick('수량', lambda h: ('수량' in h and '금액' not in h) or '주문수량' in h or '구매수량' in h)
    pick('받는사람', lambda h: '수취인명' in h or '받는분' in h or '받는사람' in h
         or (('수령자' in h or '수취인' in h) and not any(x in h for x in ('전화', '휴대', '연락', '번호'))))
    pick('전화', lambda h: '휴대전화' in h or '연락처1' in h or '휴대폰' in h or '연락처' in h)
    pick('주소', lambda h: '통합배송지' in h or '배송지' in h or '주소' in h)
    pick('우편번호', lambda h: '우편번호' in h)
    pick('송장', lambda h: '송장번호' in h and '일시' not in h)
    return found


def detect_format(df):
    """헤더에 '판매채널' 또는 '옵션정보'가 있으면 스마트스토어, 아니면 제이."""
    H = [_hdr(c) for c in df.columns]
    if any('판매채널' in h for h in H) or any('옵션정보' in h for h in H):
        return 'smartstore'
    return 'jeje'


def adapt_to_jeje(df, fields):
    """스마트스토어 df → 제이제이 컬럼배치 df (매칭키=옵션정보를 '선택사항' 자리에)."""
    n = len(df)
    out = pd.DataFrame(index=range(n), columns=JEJE_HEADERS)

    def col(field):
        i = fields.get(field)
        return df.iloc[:, i] if i is not None else None

    def put(pos, series):
        if series is not None:
            out.iloc[:, pos] = list(series)

    put(2, col('전화'))        # 수령자휴대전화
    put(4, col('받는사람'))     # 수령자
    put(5, col('우편번호'))     # 우편번호
    put(6, col('주소'))        # 주소
    put(7, col('옵션'))        # 선택사항(=매칭키)
    put(8, col('상품명'))      # 상품명
    put(10, col('수량'))       # 수량
    put(12, col('송장'))       # 송장번호(있으면 보존)
    return out


def _postprocess(output_file, log=print):
    """결과 엑셀 사후 가공 (convert 미변경): 상품명*수량 표시만.
       상품명(엑셀 9열)·수량(엑셀 11열) 기준. 수량 컬럼은 손대지 않음."""
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    상품명열, 수량열 = 9, 11
    for r in range(2, ws.max_row + 1):
        상품 = ws.cell(row=r, column=상품명열).value
        if 상품 is None or str(상품).strip() == '':
            continue
        # 상품명*수량 (수동확인/이미 * 있는 경우 제외)
        s = str(상품)
        if s != '수동확인 필요' and '*' not in s:
            q = ws.cell(row=r, column=수량열).value
            try:
                q = int(q)
                ws.cell(row=r, column=상품명열, value='%s*%d' % (s, q))
            except (ValueError, TypeError):
                pass
    wb.save(output_file)


def convert_v2(input_file, mapping_file, output_path, log=print):
    """양식 자동 감지 → (스마트스토어면 어댑터) → 기존 convert() 호출 →
       사후분석(비고) + 상품명*수량 가공. (output, 사유목록) 반환."""
    df = pd.read_excel(input_file, header=0)
    fmt = detect_format(df)
    fields = detect_fields(df)
    log(f"양식 감지: {'스마트스토어' if fmt == 'smartstore' else '제이제이'}")

    if fmt == 'smartstore':
        miss = [k for k in ('상품명', '수량', '받는사람') if k not in fields]
        if miss:
            raise RuntimeError("스마트스토어 필수 컬럼 인식 실패: %s" % miss)
        adapted = adapt_to_jeje(df, fields)
        temp = output_path + '.__adapt__.xlsx'
        adapted.to_excel(temp, index=False)
        try:
            out = convert(temp, mapping_file, output_path, log=log)
            사유 = 분석_수동확인사유(temp, mapping_file, out, log=lambda *a: None)
        finally:
            try:
                os.remove(temp)
            except OSError:
                pass
    else:
        out = convert(input_file, mapping_file, output_path, log=log)
        사유 = 분석_수동확인사유(input_file, mapping_file, out, log=lambda *a: None)

    _postprocess(out, log=log)
    return out, 사유


# ═════════════════════════════════════════════════════════
# 발주서 분리 — 검증완료 변환파일 → 택배사별 발주서(출력양식 템플릿) 생성
#   큰 틀: 각 양식의 헤더(0행)+기본행(1행, 고정값)을 읽어 같은 컬럼 구조의
#          새 .xlsx로 택배사별 출력. 변수 칸만 주문 데이터로 덮어씀.
#   ※ 대신택배·대신낱개 → 대신 양식 / 양식없는 택배사 → '기타'
# ═════════════════════════════════════════════════════════

# 변환완료 파일의 소스 컬럼(openpyxl 1-indexed) — convert_v2 출력 기준
#   phone=수령자휴대전화, phone2=수령자전화번호(폴백), memo=배송메모(배송메세지)
#   prod=상품명(이미 '회사상품명*수량'), grade=택배등급(스티로폴박스N / S~E 등)
_SRC = {'recv': 5, 'phone': 3, 'phone2': 4, 'zip': 6, 'addr': 7,
        'prod': 9, 'qty': 11, 'carrier': 12, 'total': 18, 'grade': 19,
        'memo': 14}

_CJ등급 = {'S', 'A', 'B', 'C', 'D', 'E'}   # 씨제이 박스타입에 들어갈 등급

# 택배사 → 양식 매핑.
#   kind: 'xlsx'(템플릿 그대로 채움-서식보존) / 'xls'(같은 컬럼의 새 xlsx로 출력)
#   fields = {양식열(1-indexed): 소스필드}, header_row/data_row = 1-indexed
#   no_col = 일련번호(NO.) 채울 열, balhwa_col = 발화주명/보내는분 열
_대신양식 = {'file': '대신 발주서 양식.xls', 'sheet': '대신발주서', 'kind': 'xls',
            'fields': {3: 'phone', 5: 'recv', 6: 'zip', 7: 'addr',
                       8: 'grade', 10: 'qty', 13: 'total', 14: 'prod', 15: 'memo'},
            'balhwa_col': 2}
#   8:품명=스티로폼박스(택배등급), 14:제품명=제품명*수량, 15:특기사항=배송메세지

FORM_MAP = {
    '천일': {'file': '천일택배 양식.xlsx', 'sheet': 'Sheet1', 'kind': 'xlsx',
             'header_row': 1, 'data_row': 2,
             'fields': {1: 'recv', 2: 'phone', 3: 'phone', 4: 'addr',
                        5: 'grade', 6: 'prod', 7: 'qty', 10: 'memo'}},
    #          10:배송메세지=배송메모
    '대신': _대신양식,            # 대신 + 대신낱개 (한 파일)
    '대신택배': {'file': '대신택배.xlsx', 'sheet': '출하내역', 'kind': 'xlsx',
               'header_row': 4, 'data_row': 5, 'no_col': 1, 'balhwa_col': 7,
               'fields': {2: 'recv', 4: 'addr', 5: 'phone', 8: 'prod',
                          10: 'qty', 12: 'total', 13: 'memo'}},
    # 대신택배 전용 양식: 7:보내는분=발화주명, 8:품명=제품명*수량, 12:물품가격/운임
    '씨제이': {'file': 'CJ 제이제이 입력양식.xlsx', 'sheet': 'Sheet1', 'kind': 'xlsx',
               'header_row': 1, 'data_row': 2,
               'fields': {1: 'cjgrade', 3: 'prod', 5: 'qty', 7: 'recv', 8: 'zip',
                          9: 'phone', 10: 'phone', 12: 'addr', 13: 'memo'}},
    # 1:박스타입=택배등급(S~E만), 13:요구사항=배송메세지
    '원준': {'file': '원준 발주양식.xls', 'sheet': '발주내역', 'kind': 'xls',
             'fields': {3: 'recv', 4: 'addr', 5: 'phone', 7: 'qty', 10: 'prod'}},
    '위플': {'file': '위플 발주서 양식.xls', 'sheet': '엑셀업로드양식', 'kind': 'xls',
             'fields': {2: 'prod', 5: 'qty', 10: 'recv', 11: 'phone',
                        12: 'phone', 13: 'zip', 14: 'addr', 15: 'memo'}},
    #          15:배송메세지=배송메모
    '로젠': {'file': '로젠.xls', 'sheet': '발주발송관리', 'kind': 'xls',
             'fields': {1: 'recv', 3: 'addr', 4: 'phone', 6: 'qty',
                        9: 'prod', 11: 'memo'}},   # 11:배송메세지
}


def _route(carrier):
    """택배사 → 양식 키. 대신택배는 반드시 별도, 대신·대신낱개는 한 파일."""
    cs = str(carrier).strip()
    if '대신택배' in cs:          # 대신택배 → 별도 파일 (꼭 분리)
        return '대신택배'
    if '대신' in cs:              # 대신·대신낱개 → 한 파일
        return '대신'
    if cs in FORM_MAP:
        return cs
    return '기타'


def _fill_inplace(tpl, cfg, 행들, 발화주명, outp):
    """.xlsx 템플릿을 그대로 열어 서식 보존하며 데이터 행을 채운다."""
    wb = openpyxl.load_workbook(tpl)
    ws = wb[cfg['sheet']]
    hr, dr = cfg.get('header_row', 1), cfg.get('data_row', 2)
    no_col, balhwa_col = cfg.get('no_col'), cfg.get('balhwa_col')
    ncol = 1                                     # 헤더행 마지막 비지 않은 열까지
    for c in range(1, min(ws.max_column, 80) + 1):
        if ws.cell(row=hr, column=c).value is not None:
            ncol = c
    기본행 = {c: ws.cell(row=dr, column=c).value for c in range(1, ncol + 1)}
    for i, d in enumerate(행들):
        rr = dr + i
        for c in range(1, ncol + 1):            # 기본행(고정값) 복사 (NO.열 제외)
            bv = 기본행.get(c)
            if bv is not None and c != no_col:
                ws.cell(row=rr, column=c, value=bv)
        for col, field in cfg['fields'].items():
            ws.cell(row=rr, column=col, value=d.get(field))
        if balhwa_col:
            ws.cell(row=rr, column=balhwa_col, value=발화주명)
        if no_col:
            ws.cell(row=rr, column=no_col, value=i + 1)

    # 데이터보다 많이 미리 채워진 양식 행(현불/택배·NO. 등) 잔재 비우기 → 유령 행 방지
    end = min(ws.max_row, dr + len(행들) + 300)
    for rr in range(dr + len(행들), end + 1):
        for c in range(1, ncol + 1):
            try:
                ws.cell(row=rr, column=c).value = None
            except AttributeError:        # 병합셀(MergedCell)은 읽기전용 → 건너뜀
                pass
    wb.save(outp)


def _fill_rebuild(tpl, cfg, 행들, 발화주명, outp):
    """.xls 템플릿(openpyxl 못 씀) → 같은 컬럼 구조의 새 .xlsx로 출력."""
    tdf = pd.read_excel(tpl, sheet_name=cfg['sheet'], header=None)
    ncol = tdf.shape[1]
    헤더 = [tdf.iloc[0, c] for c in range(ncol)]
    기본행 = [tdf.iloc[1, c] if tdf.shape[0] > 1 else None for c in range(ncol)]
    owb = openpyxl.Workbook(); ows = owb.active
    for c in range(ncol):
        v = 헤더[c]
        ows.cell(row=1, column=c + 1, value=(None if pd.isna(v) else v))
    for i, d in enumerate(행들):
        rr = i + 2
        for c in range(ncol):                   # 기본행(고정값) 복사
            v = 기본행[c]
            ows.cell(row=rr, column=c + 1, value=(None if pd.isna(v) else v))
        for col, field in cfg['fields'].items():
            ows.cell(row=rr, column=col, value=d.get(field))
        if cfg.get('balhwa_col'):
            ows.cell(row=rr, column=cfg['balhwa_col'], value=발화주명)
    owb.save(outp)


def 발주서분리(converted_file, 발화주명, 양식폴더, out_dir, log=print):
    """변환완료 파일을 택배사별로 나눠 각 양식에 채워 개별 파일로 저장.
       반환: [(택배사, 건수, 저장경로), ...]"""
    os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.load_workbook(converted_file, data_only=True)
    ws = wb.active

    그룹 = {}
    for r in range(2, ws.max_row + 1):
        carrier = ws.cell(row=r, column=_SRC['carrier']).value
        if carrier is None or str(carrier).strip() == '':
            continue
        d = {k: ws.cell(row=r, column=c).value for k, c in _SRC.items()}
        if not (d.get('phone') and str(d['phone']).strip()):
            d['phone'] = d.get('phone2')      # 휴대전화 없으면 일반전화로 폴백
        g = d.get('grade')
        d['cjgrade'] = g if (g is not None and
                             str(g).strip().upper() in _CJ등급) else None
        그룹.setdefault(_route(carrier), []).append(d)
    wb.close()

    stamp = datetime.datetime.now().strftime('%m%d_%H%M%S')   # 시각까지 → 덮어쓰기 방지
    결과 = []
    for key, 행들 in 그룹.items():
        outp = os.path.join(out_dir, f'{key}_{stamp}.xlsx')
        if key == '기타':                       # 양식 없는 택배사 → 단순 덤프
            owb = openpyxl.Workbook(); ows = owb.active
            ows.append(['택배사', '수령자', '전화', '우편번호', '주소',
                        '상품명', '수량', '배송비합계', '택배등급'])
            for d in 행들:
                ows.append([d['carrier'], d['recv'], d['phone'], d['zip'], d['addr'],
                            d['prod'], d['qty'], d['total'], d['grade']])
            owb.save(outp)
            결과.append((key, len(행들), outp))
            continue

        cfg = FORM_MAP[key]
        tpl = os.path.join(양식폴더, cfg['file'])
        if cfg.get('kind') == 'xlsx':
            _fill_inplace(tpl, cfg, 행들, 발화주명, outp)
        else:
            _fill_rebuild(tpl, cfg, 행들, 발화주명, outp)
        결과.append((key, len(행들), outp))

    return 결과


# ─────────────────────────────────────────────────────────
# CLI 진입점 (직접 실행 시) — 폴더 자동 탐색
# ─────────────────────────────────────────────────────────
if __name__ == '__main__':
    folder = os.path.dirname(os.path.abspath(__file__))
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        input_file = arg if os.path.isabs(arg) else os.path.join(folder, arg)
    else:
        input_file = auto_detect_input(folder)
        if not input_file:
            print(f"[오류] '{folder}\\{INPUT_GLOB}' 형식의 파일이 없습니다.")
            sys.exit(1)

    mapping_file = find_mapping_file(folder)
    if not mapping_file:
        print(f"[오류] 매핑표(.xlsx)를 찾을 수 없습니다: {folder}")
        sys.exit(1)

    try:
        out = convert(input_file, mapping_file, log=print, verbose=True)
        print(f"\n저장 완료: {out}")
    except Exception as e:
        print(f"[오류] {e}")
        sys.exit(1)
